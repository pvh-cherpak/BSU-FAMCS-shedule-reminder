# %%
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import enum
import numpy as np
import requests
from os import path, mkdir, getcwd

import pickle

from web_shedule import create_schedule_html

SCRIPT_DIR = path.dirname(__file__)

# %%
# заполнит ячейки где должен быть номр аудитории поточки, если при заполнении встретится более 1 (или 0) заполненных ячеек в рамках одной поточки
# пропустит поточку и выдаст ошибку
def fill_cells_with_pot_audit(table, potochki_audit_cells):
    errors = ''
    for (row, column, length) in potochki_audit_cells:
        not_None_elem = ''
        None_counter = 0
        for x in range(column, column + length):
            if(table[row][x] == None):
                None_counter+=1
            else:
                not_None_elem = table[row][x]
        
        if(None_counter != length - 1):
            errors += 'unexpected lot of not None elem during parsing поточка: ' + str((row, column, length))  + '\n'

        for x in range(column, column + length):
            table[row][x] = not_None_elem
        
    return (table, errors)

#вернёт (левый столбец и строку) где должны быть записанны аудитории поточки, а так же длинну поточки (строка, колонка, длинна)
def find_potochki_audit_cells(merged_cells):
    rows_of_first_field = 2
    rows_of_second_field = 1

    potocki_audit = []
    merged_cells.sort()

    for i in range(1, len(merged_cells)):
        if (merged_cells[i-1][0] == merged_cells[i][0] and 
            merged_cells[i-1][1] + rows_of_first_field == merged_cells[i][1] and
            merged_cells[i-1][2]['columns'] == merged_cells[i][2]['columns'] and
            merged_cells[i-1][2]['rows'] == rows_of_first_field and
            merged_cells[i][2]['rows'] == rows_of_second_field):
                potocki_audit.append((merged_cells[i][1] + rows_of_second_field, merged_cells[i][0], merged_cells[i][2]['columns']))
    return potocki_audit


def load_and_fill_merged_cells_with_potochki(file_path):
    # Загружаем книгу
    wb = load_workbook(filename=file_path)
    ws = wb.active

    # Создаем список списков для данных
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(list(row))
    
    big_merged_cells = []
    # Заполняем объединенные ячейки и ищем поточки
    for merged_range in ws.merged_cells.ranges:
        if(merged_range.size['columns'] > 2):
            big_merged_cells.append((merged_range.min_col - 1, merged_range.min_row - 1, merged_range.size))

        value = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
        # Корректируем индексы для 0-based
        for row in range(merged_range.min_row - 1, merged_range.max_row):
            for col in range(merged_range.min_col - 1, merged_range.max_col):
                
                if row < len(data) and col < len(data[0]):
                    data[row][col] = value
    
    potocki_audit = find_potochki_audit_cells(big_merged_cells)
    data, errors = fill_cells_with_pot_audit(data, potocki_audit)

    return data, errors


# не пашет сравненние
# if __name__ == "__main__":
#     for i in range(1):
#         table, errors = load_and_fill_merged_cells_with_potochki(path.join('unit_test_data', 'test_' + str(i),'act_table.xlsx'))
#         # with open('table.pkl', 'wb') as file:
#         #     pickle.dump(table, file)
#         df1 = pd.DataFrame(table[1:], columns=table[0])  # первая строка - заголовки
#         df1.to_csv("preprocessed.csv")
#         df_expected = pd.read_csv(path.join('unit_test_data', 'test_' + str(i), 'preprocessed.csv'))
#         if (df_expected.equals(df1)):
#             print ( f'test {i}: ok')
#         else:
#             print ( f'test {i}: ERRRRO')


def check_group_pos(table):
    group_row = 10
    group_column_start = 2
    group_column_end = 33
    group_expected = ['1 группа', '1 группа', '2 группа', '2 группа', '3 группа', '3 группа', '4 группа', '4 группа', None, '5 группа', '5 группа', '6 группа', '6 группа', '7 группа', '7 группа', None, '8 группа', '8 группа', '9 группа', '9 группа', None, '10 группа', '10 группа', '11 группа', '11 группа', '12 группа', '12 группа', '13 группа', '13 группа', '14 группа (иностр. Ст)', '14 группа (иностр. Ст)']

    if(table[10][2:33] == group_expected):
        return True
    return False


# %%
class DayOfWeek(enum.Enum):
    monday = (15, 6)
    tuesday = (40, 5)
    wednesday = (61, 6)
    thursday = (86, 6)
    friday = (111, 5)
    saturday = (132, 4)

class Group(enum.Enum):
    group_1 = 2
    group_2 = 4
    group_3 = 6
    group_4 = 8
    group_5 = 11
    group_6 = 13
    group_7 = 15
    group_8 = 18
    group_9 = 20
    group_10 = 23
    group_11 = 25
    group_12 = 27
    group_13 = 29
    group_14 = 31


def parse_day_at_group(table, group: Group, day: DayOfWeek):
    start_row, lot_of_pair = day.value
    group_column = group.value
    
    day_shedul = []
    for para_number in range(lot_of_pair):
        para_decription = []
        for offset in range(para_number * 4, (para_number + 1) * 4):
            row = start_row + offset
            para_decription.append(str(table[row][group_column]))
            if(table[row][group_column] != table[row][group_column + 1]):
                para_decription[-1] += ' | ' + str(table[row][group_column + 1])

        day_shedul.append(' \n '.join(list(dict.fromkeys(para_decription))))
    return day_shedul

def parse_schedule(table):
    schedule = {}
    for group in Group:
        group_schedule = {}
        for day in DayOfWeek:
            group_schedule[day.name] = parse_day_at_group(table, group, day)
        
        schedule[group.name] = group_schedule
    return schedule

if __name__ == "__main__":
    test_dir = path.join(SCRIPT_DIR, 'unit_test_data')
    for i in range(1):
        with open(path.join(test_dir, 'test_' + str(i), 'table.pkl') , 'rb') as file:
            table = pickle.load(file) 
        schedule = parse_schedule(table)

        expexted_shedule = None
        with open(path.join(test_dir, 'test_' + str(i), 'schedule.pkl'), 'rb') as file:
            expexted_shedule = pickle.load(file)
        
        if (expexted_shedule == schedule):
            print ( f'test {i}: ok')
        else:
            print (f"test {i}: ERRRRO")
        

# %%
def download_exel(destination_path_with_name,):
    url = 'https://docs.google.com/spreadsheets/d/1zr_2iHzixbm0uwKUZuDqJPDzA5PJ-ltZ/export?format=xlsx'
    response = requests.get(url)

    if(response.status_code != 200):
        return response.status_code
    
    with open(destination_path_with_name, 'wb') as file:
        file.write(response.content)
    
    return 200


# %%
# возвращаю кортедж (вердикт, словарь групп содержащий словарь дней содержащий словарь с указанием перехода из shedule1 в shedule2 ключи это индексы пар)
_change_posl_ = "\n|\n|\n\\/\n"
_no_para_siquince_ = str(None)
def compare_two_shedul(shedule1 : dict[str, dict[str, list]], shedule2 : dict[str, dict[str, list]]):
    if(shedule1.keys() != shedule2.keys()):
        return ("ERROR: Попытка сравнения расписаний разных групп", None)
    
    answer : dict[str, dict[str, dict[int, str]]] = {}
    for group_key in shedule1.keys():
        if(shedule1[group_key].keys() != shedule2[group_key].keys()):
            return ("ERROR: Что-то с днями не то", None)
        
        answer[group_key] = {}

        for day_key in shedule1[group_key].keys():
            answer[group_key][day_key] = {}

            len1 = len(shedule1[group_key][day_key])
            len2 = len(shedule2[group_key][day_key])
            min_len = min(len1, len2)
            for i in range(min_len):
                if(shedule1[group_key][day_key][i] != shedule2[group_key][day_key][i]):
                    answer[group_key][day_key][i] = shedule1[group_key][day_key][i] + _change_posl_ + shedule2[group_key][day_key][i]
            
            if (len1 > len2):
                for i in range(min_len, len1):
                    if(shedule1[group_key][day_key][i] != _no_para_siquince_):
                        answer[group_key][day_key][i] = shedule1[group_key][day_key][i] + _change_posl_ + _no_para_siquince_
            elif (len1 < len2):
                for i in range(min_len, len2):
                    if(shedule2[group_key][day_key][i] != _no_para_siquince_):
                        answer[group_key][day_key][i] = _no_para_siquince_ + _change_posl_ + shedule2[group_key][day_key][i]

    return ("OK", answer)

if __name__ == "__main__":
    print("compare_two_shedul test")
    def word_un(x:str):
        global _no_para_siquince_
        global _change_posl_
        return x.replace(_no_para_siquince_, "None").replace(_change_posl_, "\n|\n|\n\\/\n")

    s1 = {"g1" : {"p" : ["1"], "v" : ['1', '2'], "s" : ['1', '2', '3']}}
    s2 = {"g1" : {"p" : ["1"], "v" : ['1', '2'], "s" : ['3', '2', '3']}}
    ans_s1_s2 = str(('OK', {'g1': {'p': {}, 'v': {}, 's': {0: '1\n|\n|\n\\/\n3'}}}))
    if (ans_s1_s2 == word_un(str(compare_two_shedul(s1, s2)))):
        print("test 1: success")
    else: 
        print("test 1: ERRRRROR")

    s3 = {"g1" : {"p" : ["1"], "v" : ['1', '2', '4'], "s" : ['1', 'None', '3']}}
    ans_s1_s3 = str(('OK', {'g1': {'p': {}, 'v': {2: 'None\n|\n|\n\\/\n4'}, 's': {1: '2\n|\n|\n\\/\nNone'}}}))
    if(word_un(str(compare_two_shedul(s1, s3))) == ans_s1_s3):
        print("test 2: success")
    else: 
        print("test 2: ERRRRROR")

# %%
#act_time будет использоваться в названии папки в которой будут храниться данные парсинга
def get_new_shedule(work_dir_path:str, act_time:str):
    act_dir = path.join(SCRIPT_DIR, work_dir_path, act_time)
    exel_tabel_path = path.join(act_dir, "google_table.xlsx")
    log_path = path.join(act_dir, "log.txt")
    preprocessed_table_path = path.join(act_dir, "preprocessed.csv")

    if(path.isdir(act_dir)):
        return ("In work folder there is folder with given time", None)
    mkdir(act_dir)

    HTTP_status_code = download_exel(exel_tabel_path)

    if(HTTP_status_code != 200):
        return (f"HTTP request falture. Response code {HTTP_status_code} ", None)
    
    table, errors = load_and_fill_merged_cells_with_potochki(exel_tabel_path)

    with open(log_path, 'w') as file:
        file.write(errors) 
    pd.DataFrame(table[1:], columns=table[0]).to_csv(preprocessed_table_path)


    if(check_group_pos(table)):
        return ("ERROR: group row was not found", None)
    
    schedule = parse_schedule(table)

    html_content = create_schedule_html(schedule)
    html_shedule_path = path.join(act_dir, 'schedule.html')
    with open(html_shedule_path, 'w', encoding='utf-8') as f:
        f.write(html_content)
        
    return ("OK", schedule)


# get_new_shedule("parser_work_folder", "123")
if __name__ == '__main__':
    print(get_new_shedule("parser_work_folder", "567"))